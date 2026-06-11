from datetime import datetime
from pathlib import Path
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment


COLUMNS = [
    ("ID Interne", "id_interne"),
    ("Référence", "reference"),
    ("Titre", "titre"),
    ("Prix", "prix"),
    ("Description", "description"),
    ("Localisation", "localisation"),
    ("Activité", "activite"),
    ("Date Publication", "date_publication"),
    ("Photos (URLs)", "photos"),
    ("URL Source", "url"),
]

HEADER_FILL = PatternFill("solid", fgColor="1F4E79")
HEADER_FONT = Font(bold=True, color="FFFFFF")


def export_to_excel(listings: list[dict], output_dir: str = ".") -> str:
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Annonces"

    # En-têtes
    for col_idx, (label, _) in enumerate(COLUMNS, 1):
        cell = ws.cell(row=1, column=col_idx, value=label)
        cell.fill = HEADER_FILL
        cell.font = HEADER_FONT
        cell.alignment = Alignment(horizontal="center")

    # Données
    for row_idx, listing in enumerate(listings, 2):
        for col_idx, (_, key) in enumerate(COLUMNS, 1):
            value = listing.get(key, "")
            if isinstance(value, list):
                value = "\n".join(value)
            ws.cell(row=row_idx, column=col_idx, value=value)

    # Largeurs de colonnes
    col_widths = [12, 12, 50, 15, 60, 20, 25, 18, 60, 60]
    for col_idx, width in enumerate(col_widths, 1):
        ws.column_dimensions[ws.cell(row=1, column=col_idx).column_letter].width = width

    timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
    filename = f"annonces_{timestamp}.xlsx"
    filepath = Path(output_dir) / filename
    wb.save(filepath)
    return str(filepath)
